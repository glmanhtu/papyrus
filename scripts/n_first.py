import argparse

import matplotlib
from matplotlib import pyplot as plt
from openpyxl.styles.fills import StopList
from scipy.cluster.hierarchy import linkage, dendrogram
import pandas as pd
from openpyxl.styles import Color, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook

# matplotlib.use('TkAgg')

parser = argparse.ArgumentParser()

parser.add_argument('--similarity_file', type=str, help='Path to the similarity file', required=True)
parser.add_argument('--n_first', type=int, default=20)
parser.add_argument('--output_file', type=str, help='Path to the output file', required=True)

args = parser.parse_args()

# Load similarity matrix from CSV file
similarity_matrix = pd.read_csv(args.similarity_file, index_col=0)
top_similarities = {}
top_similarities_value = {}
min_val, max_val = 99999., 0.
for category in similarity_matrix.columns:
    n_first = similarity_matrix[category].nlargest(args.n_first)
    top_similarities[category] = n_first.index
    top_similarities_value[category] = n_first.values
    if min_val > n_first.min():
        min_val = n_first.min()
    if max_val < n_first.max():
        max_val = n_first.max()


# Function to calculate the color based on revenue
def get_color(value):
    # Red to green gradient
    max_color = (255, 0, 0)
    min_color = (0, 255, 0)

    scaled_value = (value - min_val) / (max_val - min_val)

    r = int((max_color[0] - min_color[0]) * scaled_value + min_color[0])
    g = int((max_color[1] - min_color[1]) * scaled_value + min_color[1])
    b = int((max_color[2] - min_color[2]) * scaled_value + min_color[2])

    return PatternFill(start_color=Color(rgb=f'{r:02X}{g:02X}{b:02X}'),
                       end_color=Color(rgb=f'{r:02X}{g:02X}{b:02X}'), fill_type='solid')



# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Default"

for col_num, category in enumerate(top_similarities.keys(), 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = category
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

for idx in range(args.n_first):
    for col_num, column in enumerate(top_similarities.keys(), 1):
        cell = ws.cell(row=idx + 2, column=col_num)
        cell.value = top_similarities[column][idx]
        colour = get_color(top_similarities_value[column][idx])
        cell.fill = colour

ws = wb.create_sheet("Similarities")
for col_num, category in enumerate(top_similarities.keys(), 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = category
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

for idx in range(args.n_first):
    for col_num, column in enumerate(top_similarities.keys(), 1):
        cell = ws.cell(row=idx + 2, column=col_num)
        cell.value = top_similarities_value[column][idx]
        colour = get_color(top_similarities_value[column][idx])
        cell.fill = colour

wb.save(args.output_file)
